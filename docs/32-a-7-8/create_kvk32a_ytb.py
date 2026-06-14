"""
KVK Madde 32/A — İndirimli Kurumlar Vergisi
YTB Çalışma Kağıdı (Multi-YTB, Template Sheet Approach)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ─── COLOR PALETTE ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # Dark navy
C_HEADER_FG   = "FFFFFF"   # White
C_INPUT_BG    = "FFFDE7"   # Light yellow
C_FORMULA_BG  = "E3F2FD"   # Light blue
C_RESULT_BG   = "E8F5E9"   # Light green
C_WARNING_BG  = "FFECB3"   # Amber
C_SECTION_BG  = "D6DCE4"   # Light steel blue for section headers
C_SUBSECT_BG  = "EBF0F8"   # Very light blue for sub-section headers
C_RED_FG      = "C62828"   # Red text
C_GREEN_FG    = "1B5E20"   # Dark green text
C_BORDER      = "B0BEC5"   # Blue grey border

# Number formats
FMT_NUMBER   = '#,##0.00'
FMT_PERCENT  = '0.00%'
FMT_INT      = '#,##0'
FMT_DATE     = 'DD.MM.YYYY'

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    t = Side(style="thin", color=C_BORDER)
    n = Side(style=None)
    if sides == "all":
        return Border(left=t, right=t, top=t, bottom=t)
    if sides == "outer":
        m = Side(style="medium", color="7F7F7F")
        return Border(left=m, right=m, top=m, bottom=m)
    if sides == "bottom":
        return Border(bottom=t)
    return Border(left=t, right=t, top=t, bottom=t)

def medium_border():
    m = Side(style="medium", color="4F4F4F")
    return Border(left=m, right=m, top=m, bottom=m)

def apply_style(cell, bg=None, fg="000000", bold=False, size=10,
                h_align="left", v_align="center", wrap=False,
                italic=False, border="thin", fmt=None, num_format=None):
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = align(h_align, v_align, wrap)
    if border == "thin":
        cell.border = thin_border("all")
    elif border == "outer":
        cell.border = medium_border()
    elif border == "none":
        pass
    if num_format:
        cell.number_format = num_format


# ─── SHEET 1: KONTROL PANELİ ─────────────────────────────────────────────────

def build_kontrol_paneli(ws):
    ws.title = "KONTROL PANELİ"

    # Column widths
    col_widths = [3, 30, 20, 18, 18, 18, 18, 18, 18, 18, 5]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── TITLE ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value = "KVK Md. 32/A — Yatırıma Katkı Tutarı (YKT) ve İndirimli KV Çalışma Kağıdı"
    apply_style(c, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=14,
                h_align="center", v_align="center", border="outer")
    ws.row_dimensions[3].height = 6

    # ── MÜKELLEf BİLGİLERİ ────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.merge_cells("B4:C4")
    c = ws["B4"]
    c.value = "MÜKELLEf BİLGİLERİ"
    apply_style(c, bg=C_SECTION_BG, bold=True, size=10, h_align="center", border="thin")
    ws["D4"].border = thin_border("all")

    fields = [
        ("B5", "C5", "D5", "G5", "Mükellef Unvanı"),
        ("B6", "C6", "D6", "G6", "Vergi Kimlik Numarası (VKN)"),
        ("B7", "C7", "D7", "E7", "Dönem Yılı"),
    ]
    for label_start, label_end, val_start, val_end, label_text in fields:
        row = int(label_start[1:])
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"{label_start}:{label_end}")
        c = ws[label_start]
        c.value = label_text
        apply_style(c, bg=C_SUBSECT_BG, bold=True, size=10, border="thin")
        ws.merge_cells(f"{val_start}:{val_end}")
        c2 = ws[val_start]
        apply_style(c2, bg=C_INPUT_BG, border="thin")

    ws.row_dimensions[8].height = 6

    # ── ÖN KONTROL TABLOSU ────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 20
    ws.merge_cells("B9:G9")
    c = ws["B9"]
    c.value = "ÖN KONTROL TABLOSU"
    apply_style(c, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=11,
                h_align="center", border="thin")

    ws.row_dimensions[10].height = 18
    headers_on = ["B10", "C10", "D10", "E10", "F10", "G10"]
    header_texts = ["No", "Kontrol Konusu", "", "", "", "Durum (EVET/HAYIR)"]
    col_merges_h = [None, ("C10", "F10"), None, None, None, None]
    for i, (cell_addr, txt) in enumerate(zip(headers_on, header_texts)):
        pass
    ws.merge_cells("C10:F10")
    for addr, txt in [("B10", "No"), ("C10", "Kontrol Konusu"),
                      ("G10", "Durum")]:
        c = ws[addr]
        c.value = txt
        apply_style(c, bg=C_SECTION_BG, bold=True, h_align="center", border="thin")

    on_kontrol_items = [
        "YTB(ler) geçerli ve iptal/revize edilmemiş",
        "Finans/sigortacılık/adi ortaklık kapsamı dışı",
        "Her YTB için kazanç ayrıştırması yapıldı",
        "Çapraz YTB YKT uygulanmadı",
        "YİAKV değerlendirmesi yapıldı",
        "YMM raporu yükümlülüğü kontrol edildi (49 No'lu Sirkül)",
    ]
    dv_evet_hayir = DataValidation(
        type="list", formula1='"EVET,HAYIR"', allow_blank=True,
        showErrorMessage=True, error="EVET veya HAYIR giriniz",
        errorTitle="Geçersiz Değer"
    )
    ws.add_data_validation(dv_evet_hayir)

    for i, item in enumerate(on_kontrol_items, 1):
        r = 10 + i
        ws.row_dimensions[r].height = 18
        c_no = ws.cell(row=r, column=2, value=str(i))
        apply_style(c_no, bg="FFFFFF", h_align="center", border="thin")
        ws.merge_cells(f"C{r}:F{r}")
        c_item = ws[f"C{r}"]
        c_item.value = item
        apply_style(c_item, bg="FFFFFF", border="thin")
        c_status = ws[f"G{r}"]
        apply_style(c_status, bg=C_INPUT_BG, h_align="center", border="thin")
        dv_evet_hayir.add(c_status)

    ws.row_dimensions[17].height = 6

    # ── YTB ÖZET TABLOSU ──────────────────────────────────────────────────────
    ws.row_dimensions[18].height = 20
    ws.merge_cells("B18:J18")
    c = ws["B18"]
    c.value = "YTB ÖZET TABLOSU (Tüm YTB'lerin Özeti)"
    apply_style(c, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=11,
                h_align="center", border="thin")

    ws.row_dimensions[19].height = 36
    ozet_headers = [
        "YTB No", "Durum", "Yatırım Tutarı\n(TL)", "YKT\n(TL)",
        "Önceki Kümülatif\nKullanım (TL)", "Bu Dönem\nKullanım (TL)",
        "Bu Dönem KV\nTasarrufu (TL)", "Kalan YKT\n(TL)"
    ]
    ozet_cols = list("BCDEFGHIJ")
    for col_letter, header in zip(ozet_cols[:len(ozet_headers)], ozet_headers):
        c = ws[f"{col_letter}19"]
        c.value = header
        apply_style(c, bg=C_SECTION_BG, bold=True, h_align="center",
                    v_align="center", wrap=True, border="thin")

    # 10 YTB rows
    ytb_sheet_names = [f"YTB-{i}" for i in range(1, 11)]
    for i in range(1, 11):
        r = 19 + i
        ws.row_dimensions[r].height = 18
        sn = f"'YTB-{i}'"
        # YTB No
        c = ws.cell(row=r, column=2)
        c.value = f"=IF(ISNUMBER(MATCH({sn}!A1,{sn}!A1,0)),{sn}!A3,\"YTB-{i}\")"
        # Simpler: just reference the sheet
        ws.cell(row=r, column=2).value = f"=IFERROR({sn}!A3,\"YTB-{i}\")"
        apply_style(ws.cell(row=r, column=2), bg=C_FORMULA_BG, h_align="center", border="thin")
        # Durum (A.11) -> A-col relay A4
        ws.cell(row=r, column=3).value = f"=IFERROR({sn}!A4,\"\")"
        apply_style(ws.cell(row=r, column=3), bg=C_FORMULA_BG, h_align="center", border="thin")
        # Yatırım Tutarı (A.5) -> A-col relay A5
        ws.cell(row=r, column=4).value = f"=IFERROR({sn}!A5,0)"
        apply_style(ws.cell(row=r, column=4), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)
        # YKT Tutarı (A.7) -> A-col relay A6
        ws.cell(row=r, column=5).value = f"=IFERROR({sn}!A6,0)"
        apply_style(ws.cell(row=r, column=5), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)
        # Önceki Kümülatif Kullanım (F.2) -> A-col relay A7
        ws.cell(row=r, column=6).value = f"=IFERROR({sn}!A7,0)"
        apply_style(ws.cell(row=r, column=6), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)
        # Bu Dönem Kullanım (F.8) -> A-col relay A8
        ws.cell(row=r, column=7).value = f"=IFERROR({sn}!A8,0)"
        apply_style(ws.cell(row=r, column=7), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)
        # KV Tasarrufu (F.7) -> A-col relay A9
        ws.cell(row=r, column=8).value = f"=IFERROR({sn}!A9,0)"
        apply_style(ws.cell(row=r, column=8), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)
        # Kalan YKT (F.10) -> A-col relay A10
        ws.cell(row=r, column=9).value = f"=IFERROR({sn}!A10,0)"
        apply_style(ws.cell(row=r, column=9), bg=C_FORMULA_BG, h_align="right", border="thin",
                    num_format=FMT_NUMBER)

    # TOPLAM row
    r_total = 30
    ws.row_dimensions[r_total].height = 20
    ws.cell(row=r_total, column=2).value = "TOPLAM"
    apply_style(ws.cell(row=r_total, column=2), bg=C_RESULT_BG, bold=True,
                h_align="center", border="thin")
    ws.cell(row=r_total, column=3).value = ""
    apply_style(ws.cell(row=r_total, column=3), bg=C_RESULT_BG, border="thin")
    for col_idx in range(4, 10):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=r_total, column=col_idx).value = \
            f"=SUM({col_letter}20:{col_letter}29)"
        apply_style(ws.cell(row=r_total, column=col_idx),
                    bg=C_RESULT_BG, bold=True, h_align="right",
                    border="thin", num_format=FMT_NUMBER)

    ws.row_dimensions[31].height = 6

    # ── FINAL RESULTS BOX ─────────────────────────────────────────────────────
    ws.row_dimensions[32].height = 20
    ws.merge_cells("B32:J32")
    c = ws["B32"]
    c.value = "DÖNEM SONU BEYANNAME SONUÇLARI"
    apply_style(c, bg="2E7D32", fg=C_HEADER_FG, bold=True, size=11,
                h_align="center", border="outer")

    final_rows = [
        (33, "Toplam Bu Dönem KV Tasarrufu (Beyanname 710/720)", "=H30"),
        (34, "Tüm YTB'ler Kalan YKT Toplamı (Devreden)", "=J30"),
    ]
    for r, label, formula in final_rows:
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"B{r}:H{r}")
        c_lbl = ws[f"B{r}"]
        c_lbl.value = label
        apply_style(c_lbl, bg=C_RESULT_BG, bold=True, border="thin")
        ws.merge_cells(f"I{r}:J{r}")
        c_val = ws[f"I{r}"]
        c_val.value = formula
        apply_style(c_val, bg=C_RESULT_BG, bold=True, fg=C_GREEN_FG,
                    h_align="right", border="thin", num_format=FMT_NUMBER)

    ws.row_dimensions[35].height = 6

    # ── INSTRUCTIONS ─────────────────────────────────────────────────────────
    ws.row_dimensions[36].height = 18
    ws.merge_cells("B36:J36")
    c = ws["B36"]
    c.value = ("ℹ️  KULLANIM: Her YTB için ayrı bir YTB-X sayfası kullanın. "
               "Daha fazla YTB için mevcut bir YTB sayfasını kopyalayıp "
               "adını değiştirin (YTB-6, YTB-7, ...) ve formülleri güncelleyin.")
    apply_style(c, bg=C_WARNING_BG, italic=True, size=9,
                h_align="left", border="thin", wrap=True)
    ws.row_dimensions[36].height = 30


# ─── YTB SHEET BUILDER ───────────────────────────────────────────────────────

def build_ytb_sheet(ws, ytb_num=1):
    """Build a complete YTB working paper sheet."""
    ws.title = f"YTB-{ytb_num}"

    # Column widths: A(no col), B(label), C(value), D(note)
    col_widths = [3, 52, 22, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1  # current row tracker

    def new_row(height=18):
        nonlocal row
        ws.row_dimensions[row].height = height
        row += 1
        return row - 1

    def title_row(text, bg=C_HEADER_BG, fg=C_HEADER_FG, size=12, height=30):
        r = new_row(height)
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = text
        apply_style(c, bg=bg, fg=fg, bold=True, size=size,
                    h_align="center", v_align="center", border="outer")
        return r

    def section_header(text, bg=C_SECTION_BG, height=22):
        r = new_row(height)
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = text
        apply_style(c, bg=bg, bold=True, size=10,
                    h_align="left", v_align="center", border="thin")
        return r

    def input_row(label, note="", height=18, num_fmt=None):
        r = new_row(height)
        c_lbl = ws[f"B{r}"]
        c_lbl.value = label
        apply_style(c_lbl, bg="FFFFFF", border="thin")
        c_val = ws[f"C{r}"]
        apply_style(c_val, bg=C_INPUT_BG, border="thin",
                    num_format=num_fmt if num_fmt else "@")
        c_note = ws[f"D{r}"]
        c_note.value = note
        apply_style(c_note, fg="5D4037", italic=True, size=9,
                    border="thin", wrap=True)
        return r

    def formula_row(label, formula, note="", height=18,
                    num_fmt=FMT_NUMBER, bg=C_FORMULA_BG, bold=False):
        r = new_row(height)
        c_lbl = ws[f"B{r}"]
        c_lbl.value = label
        apply_style(c_lbl, bg="F5F5F5", border="thin", bold=bold)
        c_val = ws[f"C{r}"]
        c_val.value = formula
        apply_style(c_val, bg=bg, border="thin", h_align="right",
                    num_format=num_fmt, bold=bold)
        c_note = ws[f"D{r}"]
        c_note.value = note
        apply_style(c_note, fg="5D4037", italic=True, size=9,
                    border="thin", wrap=True)
        return r

    def result_row(label, formula, note="", height=20, num_fmt=FMT_NUMBER):
        r = new_row(height)
        c_lbl = ws[f"B{r}"]
        c_lbl.value = label
        apply_style(c_lbl, bg=C_RESULT_BG, bold=True, border="thin")
        c_val = ws[f"C{r}"]
        c_val.value = formula
        apply_style(c_val, bg=C_RESULT_BG, bold=True, fg=C_GREEN_FG,
                    border="thin", h_align="right", num_format=num_fmt)
        c_note = ws[f"D{r}"]
        c_note.value = note
        apply_style(c_note, fg="5D4037", italic=True, size=9,
                    border="thin", wrap=True)
        return r

    def spacer(height=6):
        r = new_row(height)
        for col in "ABCD":
            ws[f"{col}{r}"].fill = fill("FFFFFF")
        return r

    # ── MAIN TITLE ─────────────────────────────────────────────────────────────
    spacer(4)
    title_row(f"YTB-{ytb_num}  —  KVK Md.32/A İndirimli KV Çalışma Kağıdı", height=34)
    spacer(4)

    # ── SECTION A0: ÖN KOŞUL KONTROLLERİ ─────────────────────────────────────
    section_header("BÖLÜM A0 — ÖN KOŞUL KONTROLLERİ", bg=C_SECTION_BG)

    dv_yn = DataValidation(type="list", formula1='"EVET,HAYIR"',
                           allow_blank=True)
    ws.add_data_validation(dv_yn)

    a0_items = [
        ("A0.1", "YTB geçerli mi? (İptal veya revize edilmemiş)"),
        ("A0.2", "Sektör uygun mu? (Finans/sigorta/adi ortaklık değil)"),
        ("A0.3", "Bu YTB'den elde edilen kazanç bu YTB'ye mi uygulanıyor?"),
        ("A0.4", "Yatırım/işletme dönemi doğru tespit edildi mi?"),
        ("A0.5", "Önceki dönem YKT kümülatif kullanımı doğru girildi mi?"),
        ("A0.6", "Çapraz YTB YKT uygulaması yok mu?"),
        ("A0.7", "YİAKV uyarlaması değerlendirildi mi?"),
        ("A0.8", "YMM raporu yükümlülüğü kontrol edildi mi?"),
    ]
    a0_rows = {}
    for code, label in a0_items:
        r = new_row(18)
        ws[f"B{r}"].value = f"{code}  {label}"
        apply_style(ws[f"B{r}"], bg="FFFFFF", border="thin")
        ws[f"C{r}"].value = ""
        apply_style(ws[f"C{r}"], bg=C_INPUT_BG, h_align="center", border="thin")
        dv_yn.add(ws[f"C{r}"])
        ws[f"D{r}"].value = "EVET veya HAYIR seçin"
        apply_style(ws[f"D{r}"], fg="5D4037", italic=True, size=9, border="thin")
        a0_rows[code] = r

    # A0 SONUÇ
    r_a0_son = new_row(22)
    ws[f"B{r_a0_son}"].value = "A0.9  Ön Koşul Sonucu"
    apply_style(ws[f"B{r_a0_son}"], bold=True, bg="F5F5F5", border="thin")
    # Build formula referencing all A0 rows
    and_args = ",".join('C{}="EVET"'.format(a0_rows[k]) for k in a0_rows)
    formula_a0 = '=IF(AND({}),"✓ Hesaplanabilir","✗ Ön Koşul Eksik")'.format(and_args)
    ws[f"C{r_a0_son}"].value = formula_a0
    apply_style(ws[f"C{r_a0_son}"], bold=True, h_align="center", border="thin")
    ws[f"D{r_a0_son}"].value = "Tüm koşullar EVET ise hesaplama yapılabilir"
    apply_style(ws[f"D{r_a0_son}"], italic=True, size=9, border="thin")
    spacer()

    # ── SECTION A: YTB KİMLİK BİLGİLERİ ──────────────────────────────────────
    section_header("BÖLÜM A — YTB KİMLİK BİLGİLERİ (Girdi)")

    dv_bolge = DataValidation(type="list", formula1='"1,2,3,4,5,6"',
                              allow_blank=True)
    ws.add_data_validation(dv_bolge)
    dv_tur = DataValidation(
        type="list",
        formula1='"Komple Yeni,Tevsi,Modernizasyon,Diğer"',
        allow_blank=True)
    ws.add_data_validation(dv_tur)
    dv_donem = DataValidation(
        type="list",
        formula1='"Yatırım Döneminde,İşletme Döneminde,Tamamlandı-İşletme Devam"',
        allow_blank=True)
    ws.add_data_validation(dv_donem)

    r_a1  = input_row("A.1   YTB Numarası", "Örn: 2023/B-123456")
    r_a2  = input_row("A.2   YTB Düzenleme Tarihi", "GG.AA.YYYY formatında", num_fmt=FMT_DATE)
    r_a3  = input_row("A.3   Yatırım Bölgesi (1-6)",
                      "Bölge1→0% / Bölge2→10% / Bölge3→20% / Bölge4→30% / Bölge5→50% / Bölge6→80%")
    dv_bolge.add(ws[f"C{r_a3}"])
    r_a4  = input_row("A.4   Yatırım Türü",
                      "Komple Yeni / Tevsi / Modernizasyon / Diğer")
    dv_tur.add(ws[f"C{r_a4}"])
    r_a5  = input_row("A.5   Toplam Yatırım Tutarı (TL)",
                      "YTB'de belirtilen toplam yatırım tutarı", num_fmt=FMT_NUMBER)
    r_a6  = input_row("A.6   YKT Oranı (%)",
                      "YTB'den gelen yatırıma katkı oranı", num_fmt="0.00")
    r_a7  = formula_row("A.7   YKT Tutarı (TL)",
                        f"=IF(C{r_a5}*C{r_a6}/100=0,0,C{r_a5}*C{r_a6}/100)",
                        "= A.5 × A.6 / 100",
                        num_fmt=FMT_NUMBER)
    r_a8  = input_row("A.8   Vergi İndirim Oranı (%)",
                      "YTB'den gelen vergi indirimi oranı", num_fmt="0.00")
    r_a9  = input_row("A.9   Normal KV Oranı (%)",
                      "Varsayılan 25 (2024+); daha önce 20", num_fmt="0.00")
    ws[f"C{r_a9}"].value = 25
    r_a10 = formula_row("A.10  İndirimli KV Oranı (%)",
                        f"=C{r_a9}*(1-C{r_a8}/100)",
                        "= A.9 × (1 - A.8/100)",
                        num_fmt="0.00")
    r_a11 = input_row("A.11  Dönem Durumu",
                      "Yatırım / İşletme / Tamamlandı-İşletme Devam")
    dv_donem.add(ws[f"C{r_a11}"])
    r_a12 = input_row("A.12  Tamamlama Vizesi Tarihi (Opsiyonel)",
                      "İşletme dönemine geçişi belgeleyiniz", num_fmt=FMT_DATE)
    spacer()

    # Store key cell references for formulas below
    # We'll use named positions to write absolute row refs

    # ── SECTION B: YATIRIM DÖNEMİ DİĞER FAALİYET KAZANCI ─────────────────────
    section_header("BÖLÜM B — YATIRIM DÖNEMİNDE DİĞER FAALİYET KAZANCI İNDİRİMİ",
                   bg=C_SUBSECT_BG)
    r_b_note = new_row(20)
    ws.merge_cells(f"A{r_b_note}:D{r_b_note}")
    c = ws[f"A{r_b_note}"]
    c.value = (f'⚠  Bu bölüm yalnızca A.11 = "Yatırım Döneminde" ise uygulanır. '
               f'İşletme dönemindeyse Bölüm D kullanınız.')
    apply_style(c, bg=C_WARNING_BG, italic=True, size=9,
                h_align="left", wrap=True, border="thin")

    r_b1 = input_row("B.1   Dönem Başı Kümülatif Yatırım Harcaması (TL)",
                     "Önceki dönemlerde yapılan yatırım harcamaları toplamı",
                     num_fmt=FMT_NUMBER)
    r_b2 = input_row("B.2   Bu Dönem Yatırım Harcaması (TL)",
                     "Cari dönemde gerçekleşen yatırım harcaması",
                     num_fmt=FMT_NUMBER)
    r_b3 = formula_row("B.3   Kümülatif Toplam Harcama (TL)",
                       f"=C{r_b1}+C{r_b2}",
                       "= B.1 + B.2", num_fmt=FMT_NUMBER)
    r_b4 = formula_row("B.4   Bölge Limit Oranı (%)",
                       f'=SWITCH(C{r_a3},1,0,2,10,3,20,4,30,5,50,6,80,0)',
                       "Bölgeye göre otomatik: B1→0, B2→10, B3→20, B4→30, B5→50, B6→80",
                       num_fmt="0.00")
    r_b5 = formula_row("B.5   YKT × Bölge Limiti (TL)",
                       f"=C{r_a7}*C{r_b4}/100",
                       "= A.7 × B.4 / 100", num_fmt=FMT_NUMBER)
    r_b6 = formula_row("B.6   Azami Kullanılabilir Limit (TL)",
                       f"=MIN(C{r_b5},C{r_b3})",
                       "= MIN(B.5, B.3)", num_fmt=FMT_NUMBER)
    r_b7 = input_row("B.7   Önceki Dönemler Kümülatif YKT Kullanımı (TL)",
                     "Bir önceki yıl beyannamesinden alınır",
                     num_fmt=FMT_NUMBER)
    r_b8 = formula_row("B.8   Bu Dönem Kullanılabilir YKT (TL)",
                       f"=MAX(0,C{r_b6}-C{r_b7})",
                       "= MAX(0, B.6 - B.7)", num_fmt=FMT_NUMBER)
    r_b9 = input_row("B.9   Diğer Faaliyet Kazancı Matrahı (TL)",
                     "Yatırım dönemi diğer faaliyet kazancı",
                     num_fmt=FMT_NUMBER)
    r_b10 = formula_row("B.10  İndirimli Matrah — Yatırım Dönemi (TL)",
                        f"=IF(C{r_b9}>0,MIN(C{r_b9},IF(C{r_a9}>0,C{r_b8}/(C{r_a9}/100),0)),0)",
                        "= MIN(B.9, B.8 / (A.9/100)) — max indirimli matrah",
                        num_fmt=FMT_NUMBER)
    r_b11 = formula_row("B.11  Normal KV (TL)",
                        f"=C{r_b10}*C{r_a9}/100",
                        "= B.10 × A.9/100", num_fmt=FMT_NUMBER)
    r_b12 = formula_row("B.12  İndirimli KV (TL)",
                        f"=C{r_b10}*C{r_a10}/100",
                        "= B.10 × A.10/100", num_fmt=FMT_NUMBER)
    r_b13 = result_row("B.13  KV Tasarrufu — Yatırım Dönemi (TL)",
                       f"=C{r_b11}-C{r_b12}",
                       "= B.11 - B.12  (Bu YTB için YKT kullanımı)")
    spacer()

    # ── SECTION C: TEVSİ KAZANÇ TESPİTİ ──────────────────────────────────────
    section_header("BÖLÜM C — TEVSİ YATIRIMLARDA KAZANÇ TESPİTİ",
                   bg=C_SUBSECT_BG)
    r_c_note = new_row(20)
    ws.merge_cells(f"A{r_c_note}:D{r_c_note}")
    c = ws[f"A{r_c_note}"]
    c.value = (f'⚠  Bu bölüm yalnızca A.4 = "Tevsi" ise uygulanır.')
    apply_style(c, bg=C_WARNING_BG, italic=True, size=9,
                h_align="left", wrap=True, border="thin")

    dv_usul = DataValidation(
        type="list",
        formula1='"Gerçek Usul,Götürü - Oransal"',
        allow_blank=True)
    ws.add_data_validation(dv_usul)

    r_c1 = input_row("C.1   Kazanç Tespit Yöntemi",
                     "Gerçek Usul veya Götürü - Oransal")
    dv_usul.add(ws[f"C{r_c1}"])
    r_c2 = input_row("C.2   YTB Kapsamı ATİK Değeri — Dönem Sonu (TL)",
                     "Yeniden değerlenmiş, tevsi yatırımına ait ATİK toplamı",
                     num_fmt=FMT_NUMBER)
    r_c3 = input_row("C.3   Toplam ATİK Değeri — Dönem Sonu (TL)",
                     "Tüm faaliyetlere ait ATİK toplamı",
                     num_fmt=FMT_NUMBER)
    r_c4 = formula_row("C.4   Götürü Oran",
                       f"=IF(C{r_c3}>0,C{r_c2}/C{r_c3},0)",
                       "= C.2 / C.3", num_fmt="0.0000")
    r_c5 = input_row("C.5   Dönem KV Matrahı — Tüm Faaliyetler (TL)",
                     "Tüm faaliyetler için hesaplanan KV matrahı",
                     num_fmt=FMT_NUMBER)
    r_c6 = input_row("C.6   Gerçek Usul Tevsi Kazancı (TL)",
                     "Sadece gerçek usulde: doğrudan tespit edilen tevsi kazancı",
                     num_fmt=FMT_NUMBER)
    r_c7 = formula_row("C.7   Tevsi İndirimli Matrah (TL)",
                       f'=IF(C{r_c1}="Götürü - Oransal",C{r_c5}*C{r_c4},C{r_c6})',
                       '= Götürü: C.5×C.4   /   Gerçek: C.6',
                       num_fmt=FMT_NUMBER, bg=C_FORMULA_BG)
    spacer()

    # ── SECTION D: İŞLETME DÖNEMİ KAZANCI ────────────────────────────────────
    section_header("BÖLÜM D — İŞLETME DÖNEMİNDE BU YTB'DEN KAZANÇ",
                   bg=C_SUBSECT_BG)
    r_d_note = new_row(20)
    ws.merge_cells(f"A{r_d_note}:D{r_d_note}")
    c = ws[f"A{r_d_note}"]
    c.value = (f'ℹ  İşletme döneminde D.1 değerini girin. Tevsi yatırımında C.7 otomatik yansıtılır.')
    apply_style(c, bg=C_FORMULA_BG, italic=True, size=9,
                h_align="left", wrap=True, border="thin")

    r_d1 = input_row("D.1   Bu YTB'den Elde Edilen Kazanç Matrahı (TL)",
                     "Tevsi dışı → doğrudan girin. Tevsi → C.7 kullanın",
                     num_fmt=FMT_NUMBER)
    # Overwrite formula for Tevsi auto-fill note (user still manually enters, C7 is a guide)
    ws[f"D{r_d1}"].value = (
        f'Tevsi ise C.7 değerini buraya kopyalayabilirsiniz: =C{r_c7}')
    r_d2 = formula_row("D.2   Normal KV (TL)",
                       f"=C{r_d1}*C{r_a9}/100",
                       "= D.1 × A.9/100", num_fmt=FMT_NUMBER)
    r_d3 = formula_row("D.3   İndirimli KV (TL)",
                       f"=C{r_d1}*C{r_a10}/100",
                       "= D.1 × A.10/100", num_fmt=FMT_NUMBER)
    r_d4 = result_row("D.4   KV Tasarrufu — İşletme Dönemi (TL)",
                      f"=C{r_d2}-C{r_d3}",
                      "= D.2 - D.3")
    spacer()

    # ── SECTION E: YDO ENDEKSLEMESİ ──────────────────────────────────────────
    section_header("BÖLÜM E — YDO ENDEKSLEMESİ (Tamamlama Sonrası)",
                   bg=C_SUBSECT_BG)
    r_e_note = new_row(20)
    ws.merge_cells(f"A{r_e_note}:D{r_e_note}")
    c = ws[f"A{r_e_note}"]
    c.value = ("ℹ  Tamamlama vizesi alındıktan sonra kalan YKT YDO ile endekslenir. "
               "2024: %63,93 — 2025: %3,30")
    apply_style(c, bg=C_FORMULA_BG, italic=True, size=9,
                h_align="left", wrap=True, border="thin")

    r_e1 = input_row("E.1   Dönem Başı Kalan YKT (TL)",
                     "Tamamlama vizesi sonrası dönem başındaki bakiye YKT",
                     num_fmt=FMT_NUMBER)
    r_e2 = input_row("E.2   YDO Oranı (%)",
                     "2024: 63,93% — 2025: 3,30% — Hazine açıklaması izleyin",
                     num_fmt="0.00")
    r_e3 = formula_row("E.3   YDO Artışı (TL)",
                       f"=C{r_e1}*C{r_e2}/100",
                       "= E.1 × E.2/100", num_fmt=FMT_NUMBER)
    r_e4 = result_row("E.4   Endekslenmiş Dönem Başı YKT (TL)",
                      f"=C{r_e1}+C{r_e3}",
                      "= E.1 + E.3")
    spacer()

    # ── SECTION F: KÜMÜLATİF YKT TAKİBİ ─────────────────────────────────────
    section_header("BÖLÜM F — KÜMÜLATİF YKT TAKİBİ VE DÖNEM SONUCU",
                   bg=C_SECTION_BG)

    r_f1  = formula_row("F.1   Toplam YKT (TL)",
                        f"=C{r_a7}", "= A.7",
                        num_fmt=FMT_NUMBER, bold=True)
    r_f2  = input_row("F.2   Önceki Dönemler Kümülatif YKT Kullanımı (TL)",
                      "GİB beyanname / vergi dairesi kayıtlarından alınır",
                      num_fmt=FMT_NUMBER)
    r_f3  = formula_row("F.3   Dönem Başı Kalan YKT (TL)",
                        f"=MAX(0,C{r_f1}-C{r_f2})",
                        "= F.1 - F.2", num_fmt=FMT_NUMBER)
    r_f4  = formula_row("F.4   Bu Dönem Toplam İndirimli Matrah (TL)",
                        f"=C{r_b10}+C{r_d1}",
                        "= B.10 + D.1", num_fmt=FMT_NUMBER)
    r_f5  = formula_row("F.5   Bu Dönem Normal KV (TL)",
                        f"=C{r_b11}+C{r_d2}",
                        "= B.11 + D.2", num_fmt=FMT_NUMBER)
    r_f6  = formula_row("F.6   Bu Dönem İndirimli KV (TL)",
                        f"=C{r_b12}+C{r_d3}",
                        "= B.12 + D.3", num_fmt=FMT_NUMBER)
    r_f7  = formula_row("F.7   Bu Dönem KV Tasarrufu (TL)",
                        f"=C{r_f5}-C{r_f6}",
                        "= F.5 - F.6", num_fmt=FMT_NUMBER,
                        bg=C_RESULT_BG, bold=True)
    r_f8  = formula_row("F.8   Bu Dönem YKT Kullanımı (TL)",
                        f"=C{r_b13}+C{r_d4}",
                        "= B.13 + D.4", num_fmt=FMT_NUMBER,
                        bg=C_RESULT_BG, bold=True)
    r_f9  = formula_row("F.9   Kontrol (F.7 - F.8 = 0?)",
                        f"=ROUND(C{r_f7}-C{r_f8},2)",
                        "Sıfır (0) olmalı — aksi hâlde hata var",
                        num_fmt=FMT_NUMBER)
    # Conditional color for F9 will be approximated via note
    ws[f"D{r_f9}"].value = "⚠  0 olmalı. Farklıysa kontrol edin."

    r_f10 = result_row("F.10  Dönem Sonu Kalan YKT (TL) — DEVREDEN",
                       f"=MAX(0,C{r_f3}-C{r_f8})",
                       "= MAX(0, F.3 - F.8)  → Bir sonraki döneme devreder",
                       height=22)
    r_f11 = formula_row("F.11  Kümülatif Kullanım Oranı (%)",
                        f"=IF(C{r_f1}>0,(C{r_f2}+C{r_f8})/C{r_f1}*100,0)",
                        "= (F.2 + F.8) / F.1 × 100",
                        num_fmt="0.00")
    # F.12 Status indicator
    r_f12 = new_row(22)
    ws[f"B{r_f12}"].value = "F.12  YKT Durum Göstergesi"
    apply_style(ws[f"B{r_f12}"], bg="F5F5F5", bold=True, border="thin")
    ws[f"C{r_f12}"].value = (
        f'=IF(C{r_a7}=0,"YKT Giriş Eksik",'
        f'IF(C{r_f10}<=0,"YKT Tükendi",'
        f'IF(C{r_f8}=0,"Bu Dönem Kullanım Yok",'
        f'IF(C{r_f11}>=95,"Son Dönem","Devam Ediyor"))))'
    )
    apply_style(ws[f"C{r_f12}"], bg=C_WARNING_BG, bold=True,
                h_align="center", border="thin")
    ws[f"D{r_f12}"].value = ("YKT Giriş Eksik / YKT Tükendi / "
                              "Bu Dönem Kullanım Yok / Son Dönem / Devam Ediyor")
    apply_style(ws[f"D{r_f12}"], italic=True, size=9, border="thin")
    spacer()

    # ── SECTION G: SONUÇ ÖZETİ ────────────────────────────────────────────────
    section_header("BÖLÜM G — SONUÇ ÖZETİ", bg="2E7D32")
    ws[f"A{row-1}"].font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")

    g_items = [
        ("Bu YTB KV Tasarrufu (TL)", f"=C{r_f7}"),
        ("Kalan YKT — Devreden (TL)", f"=C{r_f10}"),
        ("YKT Kullanım Durumu", f"=C{r_f12}"),
    ]
    for label, formula in g_items:
        r = new_row(22)
        ws[f"B{r}"].value = label
        apply_style(ws[f"B{r}"], bg=C_RESULT_BG, bold=True, border="thin")
        ws[f"C{r}"].value = formula
        apply_style(ws[f"C{r}"], bg=C_RESULT_BG, bold=True, fg=C_GREEN_FG,
                    h_align="right", border="thin", num_format=FMT_NUMBER)
        ws.merge_cells(f"D{r}:D{r}")
        ws[f"D{r}"].border = thin_border("all")
        ws[f"D{r}"].fill = fill(C_RESULT_BG)

    spacer()

    # Store critical row references on the sheet for KONTROL PANELİ formulas
    # We put YTB No in A3 (after spacer at row1, title at row2, spacer at row3)
    # The actual row assignments depend on the build. Let's track them via
    # a hidden row at the very top for easy cross-sheet reference.
    # We use rows A1:A2 as metadata (hidden, not styled)
    ws["A1"].value = f"YTB-{ytb_num}"  # sheet identifier
    ws["A3"].value = None  # Will be set below — YTB No from A.1

    # Link A3 to A.1 (YTB Numarası) for cross-sheet reference
    ws[f"A3"].value = f"=C{r_a1}"
    ws[f"A3"].font = Font(color="FFFFFF", size=8, name="Calibri")  # invisible

    # Store key refs for KONTROL PANELİ in hidden cells A4..A12
    key_refs = {
        "A4": r_a11,   # Dönem Durumu
        "A5": r_a5,    # Yatırım Tutarı
        "A6": r_a7,    # YKT Tutarı
        "A7": r_f2,    # Önceki Kullanım (F.2)
        "A8": r_f8,    # Bu Dönem Kullanım (F.8)
        "A9": r_f7,    # KV Tasarrufu (F.7)
        "A10": r_f10,  # Kalan YKT (F.10)
    }
    for cell_addr, target_row in key_refs.items():
        ws[cell_addr].value = f"=C{target_row}"
        ws[cell_addr].font = Font(color="FFFFFF", size=8, name="Calibri")

    return {
        "r_a1": r_a1, "r_a3": r_a3, "r_a4": r_a4,
        "r_a5": r_a5, "r_a7": r_a7, "r_a8": r_a8,
        "r_a9": r_a9, "r_a10": r_a10, "r_a11": r_a11,
        "r_f2": r_f2, "r_f7": r_f7, "r_f8": r_f8, "r_f10": r_f10,
    }


# ─── SHEET 7: BELGE LİSTESİ ──────────────────────────────────────────────────

def build_belge_listesi(ws):
    ws.title = "BELGE LİSTESİ"

    col_widths = [3, 48, 28, 16, 16, 32]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "KVK Md.32/A — Zorunlu Belge Listesi ve Kontrol Tablosu"
    apply_style(c, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=13,
                h_align="center", border="outer")
    ws.row_dimensions[3].height = 8

    # Column headers
    ws.row_dimensions[4].height = 30
    headers = ["Belge", "Zorunluluk", "Mevcut mu?", "Tarih / No", "Not"]
    cols = "BCDEF"
    for col, hdr in zip(cols, headers):
        c = ws[f"{col}4"]
        c.value = hdr
        apply_style(c, bg=C_SECTION_BG, bold=True, h_align="center",
                    v_align="center", border="thin")

    dv_evet = DataValidation(type="list", formula1='"EVET,HAYIR,N/A"',
                             allow_blank=True)
    ws.add_data_validation(dv_evet)

    belgeler = [
        ("YTB aslı veya noter onaylı sureti",
         "Zorunlu", "", "", "Her YTB için ayrı"),
        ("Tamamlama vizesi belgesi",
         "İşletme döneminde zorunlu", "", "", "İşletme dönemine geçişte"),
        ("Yatırım harcamaları özet tablosu",
         "Zorunlu", "", "", "Her dönem güncel"),
        ("Kazanç ayrıştırma tablosu (Tevsi için)",
         "Tevsi yatırımlarda zorunlu", "", "", "C bölümünü destekler"),
        ("YMM tasdik raporu (49 No'lu Sirkül kapsamı)",
         "Gerekiyorsa zorunlu", "", "", "Sirkül koşullarını kontrol edin"),
        ("ATİK değerleme belgesi (Tevsi götürü usul)",
         "Tevsi götürü usulde zorunlu", "", "", "Dönem sonu değerlemesi"),
        ("Önceki dönemler kümülatif YKT kullanım tablosu",
         "Zorunlu", "", "", "F.2 için kaynak belge"),
    ]

    for i, (belge, zorunluluk, mevcut, tarih, not_text) in enumerate(belgeler, 1):
        r = 4 + i
        ws.row_dimensions[r].height = 22
        ws[f"B{r}"].value = belge
        apply_style(ws[f"B{r}"], bg="FFFFFF", border="thin", wrap=True)
        ws[f"C{r}"].value = zorunluluk
        apply_style(ws[f"C{r}"], bg=C_WARNING_BG, italic=True, size=9,
                    border="thin", wrap=True)
        ws[f"D{r}"].value = ""
        apply_style(ws[f"D{r}"], bg=C_INPUT_BG, h_align="center", border="thin")
        dv_evet.add(ws[f"D{r}"])
        ws[f"E{r}"].value = ""
        apply_style(ws[f"E{r}"], bg=C_INPUT_BG, border="thin")
        ws[f"F{r}"].value = not_text
        apply_style(ws[f"F{r}"], italic=True, size=9, border="thin", wrap=True)


# ─── SHEET 8: DENETÇİ NOTLARI ────────────────────────────────────────────────

def build_denetci_notlari(ws):
    ws.title = "DENETÇİ NOTLARI"

    col_widths = [3, 20, 70, 5]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "DENETÇİ NOTLARI — KVK Md.32/A Uygulama Rehberi"
    apply_style(c, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=13,
                h_align="center", border="outer")
    ws.row_dimensions[3].height = 8

    notes = [
        ("KVK 32/A TEMEL PRENSİPLER",
         ("• YTB, Ekonomi/Sanayi/Ticaret Bakanlığı'nca verilen Yatırım Teşvik Belgesi'dir.\n"
          "• İndirimli KV, YTB'de belirtilen katkı ve indirim oranları çerçevesinde uygulanır.\n"
          "• Her YTB bağımsız değerlendirilir; YKT bir YTB'den diğerine aktarılamaz.\n"
          "• YKT tamamen kullanılıncaya veya YTB süresi doluncaya kadar devam eder.")),
        ("YATIRIM vs İŞLETME DÖNEMİ FARKI",
         ("• Yatırım Dönemi: YTB onay tarihi → Tamamlama vizesi. Bu dönemde YALNIZCA\n"
          "  diğer faaliyet kazancından (Bölüm B) indirim yapılabilir; bölge limiti uygulanır.\n"
          "• İşletme Dönemi: Tamamlama vizesinden sonra. Bu dönemde YTB kazancından\n"
          "  (Bölüm D) indirim yapılır; bölge limiti UYGULANMAZ.\n"
          "• Yatırım döneminde indirilemeyen kümülatif miktar işletme dönemine devrolur.")),
        ("TEVSİ İÇİN MATRAH TESPİTİ",
         ("• Tevsi yatırımında kazanç tespiti için iki yöntem vardır:\n"
          "  1. Gerçek Usul: İşletme içi kayıtlardan doğrudan tespit (daha güvenilir).\n"
          "  2. Götürü - Oransal: (YTB ATİK / Toplam ATİK) × KV Matrahı.\n"
          "• Yöntem seçimi tutarlı uygulanmalıdır; dönem içinde değiştirilemez.\n"
          "• ATİK değerleri dönem sonu itibarıyla, yeniden değerleme sonrası alınır.")),
        ("ÇAPRAZ YTB UYGULAMA YASAĞI",
         ("• Bir YTB'den elde edilen kazanç yalnızca o YTB için YKT indiriminde\n"
          "  kullanılabilir; başka bir YTB'nin YKT'sine mahsup edilemez.\n"
          "• Birden fazla YTB varsa her biri için bağımsız kazanç ayrıştırması şarttır.\n"
          "• Çapraz uygulama cezai tarhiyat riskine yol açar.")),
        ("YDO ENDEKSLEMESİ KURALI",
         ("• Tamamlama vizesi alındıktan sonra devreden kalan YKT, her yıl\n"
          "  Yeniden Değerleme Oranı (YDO) ile artırılır.\n"
          "• YDO: Maliye Bakanlığı her yıl Kasım/Aralık aylarında ilan eder.\n"
          "  2024 yılı: %63,93 — 2025 yılı: %3,30 (tahmini)\n"
          "• Endekslenmiş tutar beyannamede gösterilmeli ve desteklenmelidir.")),
        ("49 NO'LU SİRKÜLERE GÖRE YMM RAPORU ZORUNLULUĞU",
         ("• Belirli koşullarda (büyük ölçekli yatırım, özel bölge vb.) YMM\n"
          "  tasdik raporu zorunludur. 49 No'lu Sirküleri mutlaka inceleyin.\n"
          "• YMM raporu olmadan yapılan indirimlerin tarhiyatta reddedilme riski vardır.\n"
          "• Rapor, beyannamenin verildiği dönemde hazırlanmış olmalıdır.")),
        ("YİAKV İLE ETKİLEŞİM",
         ("• Yurt İçi Asgari Kurumlar Vergisi (YİAKV), indirimli KV uygulamasını\n"
          "  dolaylı olarak sınırlayabilir. KV, YİAKV'nin altına düşemez.\n"
          "• YİAKV hesabında indirimli KV öncesi matrah dikkate alınır.\n"
          "• İki hesaplama paralel yapılmalı ve yüksek olan beyannamede kullanılmalıdır.\n"
          "• 2024+ dönemleri için YİAKV etkisi özellikle dikkatle değerlendirilmelidir.")),
    ]

    current_row = 4
    for title, content in notes:
        ws.row_dimensions[current_row].height = 22
        ws.merge_cells(f"B{current_row}:C{current_row}")
        c = ws[f"B{current_row}"]
        c.value = title
        apply_style(c, bg=C_SECTION_BG, bold=True, size=10,
                    h_align="left", border="thin")
        current_row += 1

        lines = content.strip().split("\n")
        content_height = max(18, len(lines) * 14)
        ws.row_dimensions[current_row].height = content_height
        ws.merge_cells(f"B{current_row}:C{current_row}")
        c = ws[f"B{current_row}"]
        c.value = content.strip()
        apply_style(c, bg="FAFAFA", size=9, h_align="left",
                    v_align="top", wrap=True, border="thin")
        current_row += 1

        # spacer
        ws.row_dimensions[current_row].height = 6
        current_row += 1


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    output_path = "/home/ziyahan/declaro/docs/32-a-7-8/KVK_32A_YTB_Indirimli_KV_Calisma_Kagidi.xlsx"

    wb = Workbook()

    # Sheet 1: KONTROL PANELİ
    ws_kontrol = wb.active
    build_kontrol_paneli(ws_kontrol)

    # Sheets 2-6: YTB-1 through YTB-5
    ytb_refs = {}
    for i in range(1, 6):
        ws_ytb = wb.create_sheet(title=f"YTB-{i}")
        refs = build_ytb_sheet(ws_ytb, ytb_num=i)
        ytb_refs[i] = refs

    # Sheet 7: BELGE LİSTESİ
    ws_belge = wb.create_sheet(title="BELGE LİSTESİ")
    build_belge_listesi(ws_belge)

    # Sheet 8: DENETÇİ NOTLARI
    ws_notlar = wb.create_sheet(title="DENETÇİ NOTLARI")
    build_denetci_notlari(ws_notlar)

    # Freeze panes
    for ws in wb.worksheets:
        ws.freeze_panes = "B5"

    # Print settings
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_area = None

    wb.save(output_path)
    print(f"✓ Workbook created: {output_path}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")

    import os
    size_kb = os.path.getsize(output_path) / 1024
    print(f"  File size: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
