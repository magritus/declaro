import io
import logging

from fastapi import APIRouter, Depends, HTTPException, Path, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.auth.deps import get_current_user, verify_calisma_owner, verify_donem_owner
from app.db.models.calisma import Calisma
from app.db.models.donem import Donem
from app.db.models.kalem_verisi import KalemVerisi
from app.db.models.user import User
from app.db.session import get_db
from app.katalog.cache import get_katalog
from app.schemas.calisma import (
    BelgeDurumuGuncelle,
    CalismaResponse,
    KalemVeriGirdisi,
    KChecklistGuncelle,
    WizardFaz0Girdi,
    WizardFaz1Girdi,
    WizardFaz2Girdi,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/donem/{donem_id}/calisma", tags=["calisma"])
kalem_router = APIRouter(prefix="/calisma", tags=["calisma"])


@router.post("", response_model=CalismaResponse, status_code=201)
async def calisma_olustur(
    donem_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await verify_donem_owner(donem_id, current_user, db)

    calisma = Calisma(donem_id=donem_id)
    db.add(calisma)
    await db.commit()
    await db.refresh(calisma)
    logger.info("Calisma created: id=%d donem=%d", calisma.id, donem_id)
    return calisma


@router.get("", response_model=list[CalismaResponse])
async def calisma_listele(
    donem_id: int,
    skip: int = 0,
    limit: int = Query(default=50, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await verify_donem_owner(donem_id, current_user, db)

    result = await db.execute(
        select(Calisma).where(Calisma.donem_id == donem_id).offset(skip).limit(limit)
    )
    return result.scalars().all()


@kalem_router.get("/{calisma_id}", response_model=CalismaResponse)
async def calisma_getir(
    calisma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    return calisma


@kalem_router.put("/{calisma_id}/wizard/faz0", response_model=CalismaResponse)
async def wizard_faz0(
    calisma_id: int,
    data: WizardFaz0Girdi,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)

    calisma.ticari_kar_zarar = data.ticari_kar_zarar
    calisma.kkeg = data.kkeg
    calisma.finansman_fonu = data.finansman_fonu
    calisma.kar_mi_zarar_mi = "kar" if data.ticari_kar_zarar > 0 else "zarar"
    calisma.wizard_faz = 1

    await db.commit()
    await db.refresh(calisma)
    return calisma


@kalem_router.put("/{calisma_id}/wizard/faz1", response_model=CalismaResponse)
async def wizard_faz1(
    calisma_id: int,
    data: WizardFaz1Girdi,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)

    if calisma.wizard_faz < 1:
        raise HTTPException(status_code=400, detail="Faz 0 tamamlanmadan faz 1 baslatılamaz")

    if not any(data.secilen_kategoriler.values()):
        raise HTTPException(status_code=400, detail="En az bir kategori secilmelidir")

    mevcut = dict(calisma.wizard_cevaplari or {})
    mevcut["faz1"] = data.secilen_kategoriler
    calisma.wizard_cevaplari = mevcut
    calisma.wizard_faz = 2

    await db.commit()
    await db.refresh(calisma)
    return calisma


@kalem_router.put("/{calisma_id}/wizard/faz2", response_model=CalismaResponse)
async def wizard_faz2(
    calisma_id: int,
    data: WizardFaz2Girdi,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)

    if calisma.wizard_faz < 2:
        raise HTTPException(status_code=400, detail="Faz 1 tamamlanmadan faz 2 baslatılamaz")

    mevcut = dict(calisma.wizard_cevaplari or {})
    mevcut["faz2"] = data.kapi_soru_cevaplari
    calisma.wizard_cevaplari = mevcut
    calisma.istek_listesi = data.secilen_kalemler
    calisma.wizard_faz = 3

    await db.commit()
    await db.refresh(calisma)
    return calisma


@kalem_router.post("/{calisma_id}/tamamla", response_model=CalismaResponse)
async def calisma_tamamla(
    calisma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    calisma.tamamlandi = True
    await db.commit()
    await db.refresh(calisma)
    return calisma


@kalem_router.post("/{calisma_id}/yeniden_ac", response_model=CalismaResponse)
async def calisma_yeniden_ac(
    calisma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    calisma.tamamlandi = False
    await db.commit()
    await db.refresh(calisma)
    return calisma


@kalem_router.delete("/{calisma_id}", status_code=204)
async def calisma_sil(
    calisma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    await db.delete(calisma)
    await db.commit()
    logger.info("Calisma deleted: id=%d", calisma_id)


@kalem_router.put("/{calisma_id}/kalem/{ic_kod}/veri")
async def kalem_veri_kaydet(
    calisma_id: int,
    ic_kod: str = Path(pattern=r"^[a-z0-9_]+$", max_length=100),
    data: KalemVeriGirdisi = ...,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)

    if ic_kod not in (calisma.istek_listesi or []):
        raise HTTPException(status_code=400, detail=f"{ic_kod} bu calismanin istek listesinde degil")

    result = await db.execute(
        select(KalemVerisi).where(KalemVerisi.calisma_id == calisma_id, KalemVerisi.ic_kod == ic_kod)
    )
    kalem_verisi = result.scalar_one_or_none()

    if kalem_verisi:
        kalem_verisi.girdi_verileri = data.girdi_verileri
    else:
        kalem_verisi = KalemVerisi(calisma_id=calisma_id, ic_kod=ic_kod, girdi_verileri=data.girdi_verileri)
        db.add(kalem_verisi)

    await db.commit()
    await db.refresh(kalem_verisi)
    return {"ic_kod": ic_kod, "girdi_verileri": kalem_verisi.girdi_verileri}


@kalem_router.put("/{calisma_id}/kalem/{ic_kod}/checklist")
async def k_checklist_guncelle(
    calisma_id: int,
    ic_kod: str = Path(pattern=r"^[a-z0-9_]+$", max_length=100),
    data: KChecklistGuncelle = ...,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    if ic_kod not in (calisma.istek_listesi or []):
        raise HTTPException(status_code=400, detail=f"{ic_kod} bu calismanin istek listesinde degil")

    result = await db.execute(
        select(KalemVerisi).where(KalemVerisi.calisma_id == calisma_id, KalemVerisi.ic_kod == ic_kod)
    )
    kalem_verisi = result.scalar_one_or_none()
    if not kalem_verisi:
        raise HTTPException(status_code=404, detail="Kalem verisi bulunamadi")

    kalem_verisi.k_checklist_durumu = data.durum
    await db.commit()
    return {"ic_kod": ic_kod, "k_checklist_durumu": data.durum}


@kalem_router.put("/{calisma_id}/kalem/{ic_kod}/belgeler")
async def belge_durumu_guncelle(
    calisma_id: int,
    ic_kod: str = Path(pattern=r"^[a-z0-9_]+$", max_length=100),
    data: BelgeDurumuGuncelle = ...,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)
    if ic_kod not in (calisma.istek_listesi or []):
        raise HTTPException(status_code=400, detail=f"{ic_kod} bu calismanin istek listesinde degil")

    result = await db.execute(
        select(KalemVerisi).where(KalemVerisi.calisma_id == calisma_id, KalemVerisi.ic_kod == ic_kod)
    )
    kalem_verisi = result.scalar_one_or_none()
    if not kalem_verisi:
        raise HTTPException(status_code=404, detail="Kalem verisi bulunamadi")

    kalem_verisi.belge_durumu = data.durum
    await db.commit()
    return {"ic_kod": ic_kod, "belge_durumu": data.durum}


@kalem_router.get("/{calisma_id}/istek-listesi/excel")
async def istek_listesi_excel(
    calisma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    calisma = await verify_calisma_owner(calisma_id, current_user, db)

    katalog = get_katalog()
    istek_listesi: list[str] = calisma.istek_listesi or []

    wb = openpyxl.Workbook()

    # --- Sayfa 1: İstek Listesi ---
    ws1 = wb.active
    ws1.title = "Istek Listesi"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Beyanname Satiri", "Istisna/Indirim Adi", "YIAKV Etkisi", "Ana Kategori"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    ws1.row_dimensions[1].height = 20

    yiakv_map = {
        "dusulur": "Dusulur",
        "dusulmez": "Dusulmez",
        "tartismali": "Tartismali",
    }

    ana_kategori_map = {
        "ar_ge_istisna": "Ar-Ge ve Teknoloji Istisnaları",
        "diger_istisnalar": "Diger Istisnalar",
        "doviz_alacak_istisnalari": "Doviz / KKM Alacak Istisnaları",
        "egitim_saglik_istisnalari": "Egitim ve Saglik Istisnaları",
        "istirak_kazanc_istisnalari": "Istirak Kazanci Istisnaları",
        "serbest_bolge_tgb_istisnalari": "Serbest Bolge ve TGB Istisnaları",
        "varlik_satis_istisnalari": "Varlik Satis Istisnaları",
        "yurtdisi_istisnalar": "Yurt Disi Istisnaları (CVOA)",
        "arge_tasarim_indirimleri": "Ar-Ge, Tasarım ve Teknogirişim İndirimleri",
        "bagis_yardim_sponsorluk": "Bağış, Yardım ve Sponsorluk İndirimleri",
        "yatirim_tesvikleri": "Yatırım Teşvikleri ve Özel İndirimler",
        "hizmet_indirimleri": "Sağlık, Eğitim ve Hizmet İndirimleri",
    }

    for row_idx, ic_kod in enumerate(istek_listesi, start=2):
        kalem = katalog.get(ic_kod)
        if kalem is None:
            continue

        bk_kodlar = list({bk.kod for bk in kalem.beyanname_kodlari})
        beyanname_satiri = "/".join(str(k) for k in sorted(bk_kodlar)) if bk_kodlar else ic_kod
        yiakv = yiakv_map.get(
            kalem.yiakv_etkisi.value if hasattr(kalem.yiakv_etkisi, "value") else kalem.yiakv_etkisi,
            str(kalem.yiakv_etkisi),
        )
        ana_kategori_raw = kalem.ana_kategori or ""
        ana_kategori = ana_kategori_map.get(ana_kategori_raw, ana_kategori_raw)

        row_fill = alt_fill if row_idx % 2 == 0 else None

        values = [beyanname_satiri, kalem.baslik, yiakv, ana_kategori]
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = left_align if col_idx > 1 else center_align
            if row_fill:
                cell.fill = row_fill

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 28

    # --- Sayfa 2+: Her kalem için belge listesi ---
    zorunlu_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    destekleyici_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    belge_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    belge_header_font = Font(bold=True, color="FFFFFF", size=10)
    baslik_font = Font(bold=True, size=13)

    for ic_kod in istek_listesi:
        kalem = katalog.get(ic_kod)
        if kalem is None:
            continue

        if not kalem.belge_listesi:
            continue

        bk_kodlar_sorted = sorted({bk.kod for bk in kalem.beyanname_kodlari})
        sayfa_adi = str(bk_kodlar_sorted[0]) if bk_kodlar_sorted else ic_kod
        sayfa_adi = sayfa_adi[:31]

        ws = wb.create_sheet(title=sayfa_adi)

        baslik_cell = ws.cell(row=1, column=1, value=kalem.baslik)
        baslik_cell.font = baslik_font
        baslik_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        belge_headers = ["No", "Belge Adi", "Kategori", "Detay", "Temin Yeri"]
        for col_idx, bh in enumerate(belge_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=bh)
            cell.fill = belge_header_fill
            cell.font = belge_header_font
            cell.alignment = center_align
        ws.row_dimensions[2].height = 18

        for belge in kalem.belge_listesi:
            row_num = 2 + belge.no
            kategori_str = "Zorunlu" if belge.kategori.value == "zorunlu" else "Destekleyici"
            fill = zorunlu_fill if belge.kategori.value == "zorunlu" else destekleyici_fill

            row_vals = [belge.no, belge.baslik, kategori_str, belge.detay or "", belge.temin_yeri or ""]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = left_align if col_idx > 1 else center_align

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"istek-listesi-calisma-{calisma_id}.xlsx"
    headers_resp = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )
