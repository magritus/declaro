import io
import re
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.models.calisma import Calisma
from app.db.models.donem import Donem
from app.db.models.kalem_verisi import KalemVerisi
from app.db.models.mukellef import Mukellef
from app.pipeline.types import PipelineSonucu
from app.schemas.kalem import KalemSchema

# ─── Renk sabitleri ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="276221", bold=True)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
YELLOW_FONT = Font(color="9C6500", bold=True)
CURRENCY_FORMAT = '#,##0.00 "₺"'


# ─── Yardımcı fonksiyonlar ───────────────────────────────────────────────────

def _temizle_dosya_adi(name: str) -> str:
    return re.sub(r"[^\w\-_.]", "_", name)[:100]


def _header_satiri(ws, row: int, cols: list[str]) -> None:
    for col_idx, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _freeze(ws) -> None:
    ws.freeze_panes = "A2"


def _alt_satir_rengi(ws, row: int, ncols: int) -> None:
    if row % 2 == 0:
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = ALT_FILL


def _para_formatla(value: Any) -> str:
    if value is None:
        return "—"
    try:
        return f"{Decimal(str(value)):,.2f} ₺"
    except Exception:
        return str(value)


# ─── 1. Kalem xlsx ──────────────────────────────────────────────────────────

def kalem_xlsx(
    kalem: KalemSchema,
    kalem_verisi: KalemVerisi | None,
    mukellef: Mukellef,
    donem: Donem,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    girdi = kalem_verisi.girdi_verileri or {} if kalem_verisi else {}
    hesap = kalem_verisi.hesap_sonucu or {} if kalem_verisi else {}
    istisna = kalem_verisi.istisna_tutari if kalem_verisi else None
    checklist_durum = kalem_verisi.k_checklist_durumu or {} if kalem_verisi else {}
    belge_durum = kalem_verisi.belge_durumu or {} if kalem_verisi else {}

    # ── Sekme 1: Veri Girişi ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Veri Girişi")
    _header_satiri(ws1, 1, ["Alan", "Değer"])
    _freeze(ws1)
    row = 2
    for alan in kalem.hesaplama_sablonu.veri_girisi_alanlari:
        val = girdi.get(alan.id, "")
        ws1.cell(row=row, column=1, value=alan.etiket)
        ws1.cell(row=row, column=2, value=str(val) if val is not None else "")
        _alt_satir_rengi(ws1, row, 2)
        row += 1
    _auto_width(ws1)

    # ── Sekme 2: Hesaplamalar ────────────────────────────────────────────
    ws2 = wb.create_sheet("Hesaplamalar")
    row = 1
    # Başlık
    c = ws2.cell(row=row, column=1, value="Ara Sonuçlar")
    c.font = Font(bold=True, size=12)
    row += 1
    _header_satiri(ws2, row, ["Alan", "Değer"])
    _freeze(ws2)
    row += 1
    for key, val in hesap.items():
        ws2.cell(row=row, column=1, value=key)
        try:
            num_val = float(val)
            cell = ws2.cell(row=row, column=2, value=num_val)
            cell.number_format = CURRENCY_FORMAT
        except (TypeError, ValueError):
            ws2.cell(row=row, column=2, value=str(val))
        _alt_satir_rengi(ws2, row, 2)
        row += 1

    # İstisna tutarı — yeşil satır
    row += 1
    ws2.cell(row=row, column=1, value="İstisna Tutarı")
    istisna_val = float(istisna) if istisna is not None else 0.0
    cell = ws2.cell(row=row, column=2, value=istisna_val)
    cell.number_format = CURRENCY_FORMAT
    for c_idx in [1, 2]:
        ws2.cell(row=row, column=c_idx).fill = GREEN_FILL
        ws2.cell(row=row, column=c_idx).font = GREEN_FONT
    _auto_width(ws2)

    # ── Sekme 3: K-Checklist ─────────────────────────────────────────────
    ws3 = wb.create_sheet("K-Checklist")
    _header_satiri(ws3, 1, ["Kod", "Soru", "Referans", "Durum"])
    _freeze(ws3)
    row = 2
    for item in kalem.k_checklist:
        durum_val = checklist_durum.get(item.id, "—")
        ws3.cell(row=row, column=1, value=item.id)
        ws3.cell(row=row, column=2, value=item.soru)
        ws3.cell(row=row, column=3, value=item.referans or "")
        durum_cell = ws3.cell(row=row, column=4, value=durum_val)
        # Renk kodlama
        if durum_val == "uygun":
            durum_cell.fill = GREEN_FILL
            durum_cell.font = GREEN_FONT
        elif durum_val == "eksik":
            durum_cell.fill = RED_FILL
            durum_cell.font = RED_FONT
        elif durum_val == "risk":
            durum_cell.fill = YELLOW_FILL
            durum_cell.font = YELLOW_FONT
        _alt_satir_rengi(ws3, row, 3)  # Sadece ilk 3 sütun için alternatif renk
        row += 1
    _auto_width(ws3)

    # ── Sekme 4: Belge Listesi ────────────────────────────────────────────
    ws4 = wb.create_sheet("Belge Listesi")
    _header_satiri(ws4, 1, ["No", "Kategori", "Başlık", "Detay", "Durum", "Not"])
    _freeze(ws4)
    row = 2
    for belge in kalem.belge_listesi:
        belge_key = str(belge.no)
        durum_info = belge_durum.get(belge_key, {})
        durum = durum_info.get("durum", "") if isinstance(durum_info, dict) else ""
        not_val = durum_info.get("not", "") if isinstance(durum_info, dict) else ""
        ws4.cell(row=row, column=1, value=belge.no)
        ws4.cell(row=row, column=2, value=belge.kategori.value)
        ws4.cell(row=row, column=3, value=belge.baslik)
        ws4.cell(row=row, column=4, value=belge.detay or "")
        ws4.cell(row=row, column=5, value=durum)
        ws4.cell(row=row, column=6, value=not_val)
        _alt_satir_rengi(ws4, row, 6)
        row += 1
    _auto_width(ws4)

    # ── Sekme 5: Mevzuat Dayanağı ─────────────────────────────────────────
    ws5 = wb.create_sheet("Mevzuat Dayanağı")
    baslik_cell = ws5.cell(row=1, column=1, value=kalem.baslik)
    baslik_cell.font = Font(bold=True, size=14)
    row = 3
    for ref in kalem.mevzuat_dayanagi:
        ws5.cell(row=row, column=1, value=ref)
        row += 1
    _auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 2. Özet xlsx ───────────────────────────────────────────────────────────

def ozet_xlsx(
    calisma: Calisma,
    mukellef: Mukellef,
    donem: Donem,
    pipeline_sonucu: PipelineSonucu,
    kalem_verileri_map: dict[str, KalemVerisi],
    katalog: Any,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # ── Sekme 1: Özet ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Özet")
    # Üst bilgi bloğu
    meta = [
        ("Mükellef", mukellef.unvan),
        ("VKN", mukellef.vkn),
        ("Dönem", f"{donem.yil} / {donem.ceyrek}"),
        ("Tarih", str(date.today())),
    ]
    for idx, (label, val) in enumerate(meta, start=1):
        ws1.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=idx, column=2, value=val)

    header_row = len(meta) + 2
    _header_satiri(ws1, header_row, ["Adım No", "Başlık", "Tutar (₺)", "Açıklama"])
    ws1.freeze_panes = f"A{header_row + 1}"

    odenecek_baslik_parcasi = "Ödenecek"
    row = header_row + 1
    for adim in pipeline_sonucu.adimlar:
        ws1.cell(row=row, column=1, value=adim.adim_no)
        ws1.cell(row=row, column=2, value=adim.baslik)
        tutar_cell = ws1.cell(row=row, column=3, value=float(adim.sonraki_deger))
        tutar_cell.number_format = CURRENCY_FORMAT
        ws1.cell(row=row, column=4, value=adim.aciklama)

        # "Ödenecek KV" satırını vurgula
        if odenecek_baslik_parcasi in adim.baslik:
            for c_idx in range(1, 5):
                ws1.cell(row=row, column=c_idx).fill = GREEN_FILL
                ws1.cell(row=row, column=c_idx).font = GREEN_FONT
        else:
            _alt_satir_rengi(ws1, row, 4)
        row += 1

    _auto_width(ws1)

    # ── Sekme 2: Kalem İstisnaları ────────────────────────────────────────
    ws2 = wb.create_sheet("Kalem İstisnaları")
    _header_satiri(ws2, 1, ["Kalem Kodu", "Başlık", "İstisna Tutarı", "Hatalar", "Uyarılar"])
    _freeze(ws2)
    row = 2
    istek = calisma.istek_listesi or []
    for ic_kod in istek:
        kalem_schema = katalog.get(ic_kod)
        baslik = kalem_schema.baslik if kalem_schema else ic_kod
        ks = pipeline_sonucu.kalem_sonuclari.get(ic_kod)
        if ks:
            istisna_val = float(ks.istisna_tutari)
            hatalar = "; ".join(ks.hatalar) if ks.hatalar else ""
            uyarilar = "; ".join(ks.uyarilar) if ks.uyarilar else ""
        else:
            kv = kalem_verileri_map.get(ic_kod)
            istisna_val = float(kv.istisna_tutari) if kv and kv.istisna_tutari is not None else 0.0
            hatalar = ""
            uyarilar = ""

        ws2.cell(row=row, column=1, value=ic_kod)
        ws2.cell(row=row, column=2, value=baslik)
        tutar_cell = ws2.cell(row=row, column=3, value=istisna_val)
        tutar_cell.number_format = CURRENCY_FORMAT
        ws2.cell(row=row, column=4, value=hatalar)
        ws2.cell(row=row, column=5, value=uyarilar)
        _alt_satir_rengi(ws2, row, 5)
        row += 1

    _auto_width(ws2)

    # ── Sekme 3: Özet Notlar ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Özet Notlar")
    notlar = [
        "YİAKV (Yurt İçi Asgari Kurumlar Vergisi) 2025 ve sonrası dönemler için geçerlidir.",
        "YİAKV oranı %10 olarak uygulanmaktadır.",
        "",
        "Kazanç varsa indirilecek kalemler yalnızca kâr durumunda uygulanmaktadır.",
    ]
    if pipeline_sonucu.kazanc_varsa_gruplari_atlanmis:
        notlar.append(
            "UYARI: Bu çalışmada ticari zarar nedeniyle 'kazanç varsa' grubundaki kalemler "
            "hesaba katılmamıştır."
        )
    if pipeline_sonucu.yiakv_uygulanmis:
        notlar.append(
            f"BİLGİ: YİAKV ({float(pipeline_sonucu.yiakv):,.2f} ₺) hesaplanan KV'den yüksek "
            "olduğundan YİAKV uygulanmıştır."
        )

    baslik = ws3.cell(row=1, column=1, value="Özet Notlar")
    baslik.font = Font(bold=True, size=14)
    for idx, not_satiri in enumerate(notlar, start=3):
        ws3.cell(row=idx, column=1, value=not_satiri)

    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
