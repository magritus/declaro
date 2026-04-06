#!/usr/bin/env python3
"""
Extract content from all xlsx working papers in docs/ to JSON for analysis.
Usage: python3 scripts/xlsx_extractor.py
"""

import json
import os
import sys
from pathlib import Path

import openpyxl

DOCS_DIR = Path(__file__).parent.parent / "docs"
OUT_DIR = Path(__file__).parent.parent / "docs" / "extracted"


def cell_value(cell):
    """Return cell value as string or None."""
    v = cell.value
    if v is None:
        return None
    return str(v).strip() if str(v).strip() else None


def extract_sheet(ws) -> dict:
    """Extract meaningful rows from a worksheet."""
    rows = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            v = cell_value(cell)
            row_data.append(v)
        # Skip fully empty rows
        if any(v is not None for v in row_data):
            rows.append(row_data)
    return {
        "name": ws.title,
        "dimensions": ws.dimensions,
        "rows": rows,
    }


def extract_workbook(path: Path) -> dict:
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(extract_sheet(ws))
        return {
            "file": path.name,
            "sheets": sheets,
        }
    except Exception as e:
        return {"file": path.name, "error": str(e)}


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_files = sorted(DOCS_DIR.glob("*.xlsx"))
    print(f"Found {len(xlsx_files)} xlsx files")

    for path in xlsx_files:
        stem = path.stem
        out_path = OUT_DIR / f"{stem}.json"
        print(f"  Processing {path.name}...", end=" ")
        data = extract_workbook(path)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print("ok")

    print(f"\nExtracted to {OUT_DIR}")


if __name__ == "__main__":
    main()
